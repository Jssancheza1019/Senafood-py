from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.http import HttpResponse
from django.utils import timezone
from django.utils.timezone import localtime
from django.conf import settings

from .models import PQRSF
from gestion.models import Usuario
from notificaciones.models import Notificacion

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from reportlab.lib.pagesizes import landscape, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer


def get_usuario_session(request): #Helper para obtener usuario y su rol desde la sesión.
    if 'usuario_id' not in request.session:
        return None, None
    try:
        usuario = Usuario.objects.get(id_usuario=request.session['usuario_id'])
        rol = usuario.rol.nombre_rol if usuario.rol else None
        return usuario, rol
    except Usuario.DoesNotExist:
        return None, None


def lista_pqrsf_view(request):
    usuario, rol = get_usuario_session(request)
    if not usuario:
        return redirect('login')

    # Filtros
    filtro_tipo = request.GET.get('tipo', '')
    filtro_estado = request.GET.get('estado', '')
    filtro_leida = request.GET.get('leida', '')

    if rol == 'Administrador':
        pqrsf_list = PQRSF.objects.all().order_by('-create_at')
    else:
        pqrsf_list = PQRSF.objects.filter(usuario=usuario).order_by('-create_at')

    if filtro_tipo:
        pqrsf_list = pqrsf_list.filter(tipo=filtro_tipo)
    if filtro_estado:
        pqrsf_list = pqrsf_list.filter(estado=filtro_estado)
    if filtro_leida != '':
        pqrsf_list = pqrsf_list.filter(leida=(filtro_leida == '1'))

    #conteos para cards
    base = PQRSF.objects.all() if rol == 'Administrador' else PQRSF.objects.filter(usuario=usuario)
    total_leidas     = base.filter(leida=True).count()
    total_no_leidas  = base.filter(leida=False).count()
    total_pendiente  = base.filter(estado='Pendiente').count()
    total_resuelta   = base.filter(estado='Resuelta').count()
    total_en_gestion = base.filter(estado='En gestión').count()

    return render(request, 'pqrs/lista.html', {
        'pqrsf_list':      pqrsf_list,
        'rol':             rol,
        'nombre_usuario':  usuario.nombre,
        'filtro_tipo':     filtro_tipo,
        'filtro_estado':   filtro_estado,
        'filtro_leida':    filtro_leida,
        'tipos':           PQRSF.TIPO_CHOICES,
        'estados':         PQRSF.ESTADO_CHOICES,
        'total_leidas':    total_leidas,
        'total_no_leidas': total_no_leidas,
        'total_pendiente': total_pendiente,
        'total_resuelta':  total_resuelta,
        'total_en_gestion':total_en_gestion,
    })



def detalle_pqrsf_view(request, id):
    usuario, rol = get_usuario_session(request)
    if not usuario:
        return redirect('login')

    pqrsf = get_object_or_404(PQRSF, id_pqrsf=id)

    # Seguridad: cliente/vendedor solo ve las suyas
    if rol != 'Administrador' and pqrsf.usuario != usuario:
        return redirect('lista_pqrsf')

    # Al ver el detalle, marcar como leída
    if not pqrsf.leida:
        pqrsf.leida = True
        if pqrsf.estado == 'Pendiente':
            pqrsf.estado = 'Leída'
        pqrsf.save()

    # Admin puede responder y cambiar estado
    if request.method == 'POST' and rol == 'Administrador':
        pqrsf.respuesta = request.POST.get('respuesta', '')
        pqrsf.estado = request.POST.get('estado', pqrsf.estado)
        pqrsf.save()

        # ✅ Notificar al usuario dueño del PQRSF
        Notificacion.objects.create(
            usuario=pqrsf.usuario,
            mensaje=f"Tu PQRSF #{pqrsf.id_pqrsf} fue respondida. Estado: {pqrsf.estado}",
            tipo='pqrsf',
            pqrsf=pqrsf,
        )

        messages.success(request, 'PQRSF actualizada correctamente.')
        return redirect('detalle_pqrsf', id=id)

    return render(request, 'pqrs/detalle.html', {
        'pqrsf': pqrsf,
        'rol': rol,
        'nombre_usuario': usuario.nombre,
        'estados': PQRSF.ESTADO_CHOICES,
    })


def crear_pqrsf_view(request):
    usuario, rol = get_usuario_session(request)
    if not usuario:
        return redirect('login')

    if request.method == 'POST':
        tipo = request.POST.get('tipo')
        descripcion = request.POST.get('descripcion')

        if not tipo or not descripcion:
            messages.error(request, 'Todos los campos son obligatorios.')
            return render(request, 'pqrsf/crear.html', {
                'nombre_usuario': usuario.nombre,
                'rol': rol,
                'tipos': PQRSF.TIPO_CHOICES,
            })

        nueva_pqrsf = PQRSF.objects.create(
            tipo=tipo,
            descripcion=descripcion,
            estado='Pendiente',
            usuario=usuario,
            leida=False,
        )

        # Notificar a todos los administradores
        admins = Usuario.objects.filter(rol__nombre_rol='Administrador', es_activo=True)
        for admin in admins:
            Notificacion.objects.create(
                usuario=admin,
                mensaje=f"Nueva PQRSF de {usuario.nombre} {usuario.apellido}: {tipo}",
                tipo='pqrsf',
                pqrsf=nueva_pqrsf,  # ✅ esto faltaba
            )

        messages.success(request, '¡PQRSF creada correctamente!')
        return redirect('lista_pqrsf')

    return render(request, 'pqrs/crear.html', {
        'nombre_usuario': usuario.nombre,
        'rol': rol,
        'tipos': PQRSF.TIPO_CHOICES,
    })

def exportar_excel_pqrsf(request):
    usuario_session, rol = get_usuario_session(request)
    if not usuario_session:
        return redirect('login')

    # Usamos select_related para traer los datos del usuario de una sola vez
    if rol == 'Administrador':
        pqrsf_list = PQRSF.objects.select_related('usuario').all().order_by('-create_at')
    else:
        pqrsf_list = PQRSF.objects.select_related('usuario').filter(usuario=usuario_session).order_by('-create_at')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'PQRSF'

    headers = ['ID', 'Tipo', 'Descripción', 'Estado', 'Respuesta', 'Usuario', 'Leída', 'Fecha Creación']
    header_fill = PatternFill(start_color='28A745', end_color='28A745', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for row, p in enumerate(pqrsf_list, 2):
        ws.cell(row=row, column=1, value=p.id_pqrsf)
        ws.cell(row=row, column=2, value=p.tipo)
        ws.cell(row=row, column=3, value=p.descripcion)
        ws.cell(row=row, column=4, value=p.estado)
        ws.cell(row=row, column=5, value=p.respuesta or 'Sin respuesta')

        # Manejo robusto para el usuario (evita el error de la imagen anterior)
        try:
            if p.usuario:
                nombre_usuario = f"{p.usuario.nombre} {p.usuario.apellido}"
            else:
                nombre_usuario = "Usuario no asignado"
        except:
            nombre_usuario = "Error: Usuario no encontrado"
        
        ws.cell(row=row, column=6, value=nombre_usuario)
        ws.cell(row=row, column=7, value='Sí' if p.leida else 'No')
        
        # Fecha con zona horaria local (Colombia)
        fecha_local = localtime(p.create_at) if p.create_at else None
        ws.cell(row=row, column=8, value=fecha_local.strftime('%d/%m/%Y %H:%M') if fecha_local else '')

    # Auto-ajuste de columnas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Reporte_PQRSF_SenaFOOD.xlsx"'
    wb.save(response)
    return response


def exportar_pdf_pqrsf(request):
    usuario_session, rol = get_usuario_session(request)
    if not usuario_session:
        return redirect('login')
    
    # filtros de la lista ──
    filtro_tipo   = request.GET.get('tipo', '')
    filtro_estado = request.GET.get('estado', '')
    filtro_leida  = request.GET.get('leida', '')

    if rol == 'Administrador':
        pqrsf_list = PQRSF.objects.select_related('usuario').all().order_by('-create_at')
    else:
        pqrsf_list = PQRSF.objects.select_related('usuario').filter(usuario=usuario_session).order_by('-create_at')

    if filtro_tipo:
        pqrsf_list = pqrsf_list.filter(tipo=filtro_tipo)
    if filtro_estado:
        pqrsf_list = pqrsf_list.filter(estado=filtro_estado)
    if filtro_leida != '':
        pqrsf_list = pqrsf_list.filter(leida=(filtro_leida == '1'))
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="Reporte_PQRSF.pdf"'

    doc = SimpleDocTemplate(
        response,
        pagesize=landscape(A4),
        leftMargin=30, rightMargin=30,
        topMargin=30, bottomMargin=30
    )

    styles = getSampleStyleSheet()
    elements = []

    # ── Estilos ──
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Title'],
        fontSize=18,
        textColor=colors.HexColor('#28A745'),
        spaceAfter=4,
        fontName='Helvetica-Bold',
        alignment=1,
    )
    subtitle_style = ParagraphStyle(
        'Subtitle',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.HexColor('#666666'),
        alignment=1,
        spaceAfter=2,
    )
    stat_style = ParagraphStyle(
        'Stat',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.HexColor('#333333'),
        alignment=1,
    )

    # ── Header con logo ──
    logo_path = os.path.join(settings.BASE_DIR, 'gestion', 'static', 'gestion', 'img', 'logo.png')
    if os.path.exists(logo_path):
        from reportlab.platypus import Image as RLImage
        logo  = RLImage(logo_path, width=80, height=80)
        logo2 = RLImage(logo_path, width=80, height=80)
        header_data = [[logo, Paragraph('Reporte de PQRSF', title_style), logo2]]
        header_table = Table(header_data, colWidths=[90, 580, 90])
        header_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ALIGN', (0, 0), (0, 0), 'LEFT'),
            ('ALIGN', (1, 0), (1, 0), 'CENTER'),
            ('ALIGN', (2, 0), (2, 0), 'RIGHT'),
        ]))
        elements.append(header_table)
    else:
        elements.append(Paragraph('Reporte de PQRSF', title_style))

    elements.append(Paragraph('SenaFOOD - Sistema de Gestión', subtitle_style))
    elements.append(Paragraph(
        f'Generado el {localtime(timezone.now()).strftime("%d/%m/%Y a las %H:%M")} '
        f'por {request.session.get("usuario_nombre", "Administrador")}',
        subtitle_style
    ))
    elements.append(Spacer(1, 6))

    # ── Línea verde ──
    from reportlab.platypus import HRFlowable
    elements.append(HRFlowable(width="100%", thickness=2, color=colors.HexColor('#28A745')))
    elements.append(Spacer(1, 8))

    # ── Estadísticas ──
    base = PQRSF.objects.all() if rol == 'Administrador' else PQRSF.objects.filter(usuario=usuario_session)
    total        = base.count()
    pendientes   = base.filter(estado='Pendiente').count()
    en_gestion   = base.filter(estado='En gestión').count()
    resueltas    = base.filter(estado='Resuelta').count()
    no_leidas    = base.filter(leida=False).count()

    stats_data = [[
        Paragraph(f'<b>Total:</b> {total}',             stat_style),
        Paragraph(f'<b>Pendientes:</b> {pendientes}',   stat_style),
        Paragraph(f'<b>En gestión:</b> {en_gestion}',   stat_style),
        Paragraph(f'<b>Resueltas:</b> {resueltas}',     stat_style),
        Paragraph(f'<b>No leídas:</b> {no_leidas}',     stat_style),
    ]]
    stats_table = Table(stats_data, colWidths=[148, 148, 148, 148, 148])
    stats_table.setStyle(TableStyle([
        ('BACKGROUND',    (0,0), (-1,-1), colors.HexColor('#f0fff4')),
        ('GRID',          (0,0), (-1,-1), 0.5, colors.HexColor('#a8d5b5')),
        ('TOPPADDING',    (0,0), (-1,-1), 8),
        ('BOTTOMPADDING', (0,0), (-1,-1), 8),
        ('ALIGN',         (0,0), (-1,-1), 'CENTER'),
    ]))
    elements.append(stats_table)
    elements.append(Spacer(1, 12))

    # ── Tabla de datos ──
    data = [['ID', 'Tipo', 'Estado', 'Usuario', 'Descripción', 'Leída', 'Fecha']]

    for p in pqrsf_list:
        try:
            nombre_u = f"{p.usuario.nombre} {p.usuario.apellido}" if p.usuario else "Sin usuario"
        except Exception:
            nombre_u = "Usuario no encontrado"

        fecha_fila = localtime(p.create_at).strftime('%d/%m/%Y %H:%M') if p.create_at else "N/A"
        descripcion = (p.descripcion[:50] + '...') if p.descripcion and len(p.descripcion) > 50 else (p.descripcion or '')

        data.append([
            str(p.id_pqrsf),
            p.tipo,
            p.estado,
            nombre_u,
            descripcion,
            'Sí' if p.leida else 'No',
            fecha_fila,
        ])

    table = Table(data, repeatRows=1, colWidths=[30, 80, 70, 120, 200, 40, 90])
    table.setStyle(TableStyle([
        ('BACKGROUND',    (0,0), (-1,0), colors.HexColor('#28A745')),
        ('TEXTCOLOR',     (0,0), (-1,0), colors.white),
        ('FONTNAME',      (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE',      (0,0), (-1,0), 9),
        ('FONTSIZE',      (0,1), (-1,-1), 8),
        ('ALIGN',         (0,0), (-1,-1), 'CENTER'),
        ('VALIGN',        (0,0), (-1,-1), 'MIDDLE'),
        ('ROWBACKGROUNDS',(0,1), (-1,-1), [colors.white, colors.HexColor('#f0fff4')]),
        ('GRID',          (0,0), (-1,-1), 0.5, colors.HexColor('#a8d5b5')),
        ('TOPPADDING',    (0,0), (-1,-1), 6),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('LEFTPADDING',   (0,0), (-1,-1), 6),
        ('RIGHTPADDING',  (0,0), (-1,-1), 6),
    ]))
    elements.append(table)

    # ── Footer ──
    def add_footer(canvas, doc):
        canvas.saveState()
        canvas.setFont('Helvetica', 8)
        canvas.setFillColor(colors.HexColor('#888888'))
        canvas.drawString(30, 15, 'SenaFOOD - Reporte de PQRSF')
        canvas.drawRightString(landscape(A4)[0] - 30, 15, f'Página {doc.page}')
        canvas.setStrokeColor(colors.HexColor('#28A745'))
        canvas.setLineWidth(1)
        canvas.line(30, 25, landscape(A4)[0] - 30, 25)
        canvas.restoreState()

    doc.build(elements, onFirstPage=add_footer, onLaterPages=add_footer)
    return response
