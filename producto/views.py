from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.http import HttpResponse
from .models import Producto, ProveedorProducto
from proveedor.models import Proveedor
from gestion.models import Usuario
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import landscape, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from datetime import datetime
import os
from django.conf import settings
from django.utils import timezone


def get_usuario_session(request):
    if 'usuario_id' not in request.session:
        return None, None
    try:
        usuario = Usuario.objects.get(id_usuario=request.session['usuario_id'])
        rol = usuario.rol.nombre_rol if usuario.rol else None
        return usuario, rol
    except Usuario.DoesNotExist:
        return None, None


def lista_productos_view(request):
    usuario, rol = get_usuario_session(request)
    if not usuario or rol != 'Administrador':
        return redirect('login')

    filtro_categoria = request.GET.get('categoria', '')
    filtro_estado    = request.GET.get('estado', '')
    filtro_stock     = request.GET.get('stock', '')

    productos = Producto.objects.all().order_by('nombre')

    if filtro_categoria:
        productos = productos.filter(categoria=filtro_categoria)
    if filtro_estado:
        productos = productos.filter(estado=filtro_estado)

    # Convertir a lista solo al final
    productos = list(productos)

    if filtro_stock == 'bajo':
        productos = [p for p in productos if p.stock_bajo]

    categorias = Producto.objects.values_list(
        'categoria', flat=True
    ).distinct().exclude(categoria=None)

    # Conteos
    base = Producto.objects.all()
    total_activos   = base.filter(estado='activo').count()
    total_inactivos = base.filter(estado='inactivo').count()
    total_agotados  = base.filter(estado='agotado').count()
    total_vencidos  = base.filter(estado='vencido').count()

    return render(request, 'producto/lista.html', {
        'productos':        productos,
        'categorias':       categorias,
        'filtro_categoria': filtro_categoria,
        'filtro_estado':    filtro_estado,
        'filtro_stock':     filtro_stock,
        'estados':          [('activo', 'Activo'), ('inactivo', 'Inactivo'), ('agotado', 'Agotado'), ('vencido', 'Vencido')],
        'total_activos':    total_activos,
        'total_inactivos':  total_inactivos,
        'total_agotados':   total_agotados,
        'total_vencidos':   total_vencidos,
    })


def detalle_producto_view(request, id):
    usuario, rol = get_usuario_session(request)
    if not usuario or rol != 'Administrador':
        return redirect('login')

    producto = get_object_or_404(Producto, id_producto=id)
    proveedores_asignados = ProveedorProducto.objects.filter(producto=producto)
    proveedores_disponibles = Proveedor.objects.filter(es_activo=True).exclude(
    id_proveedor__in=proveedores_asignados.values_list('proveedor__id_proveedor', flat=True)
    )

    margen = None
    if producto.precio_venta and producto.costo_unitario:
        margen = producto.precio_venta - producto.costo_unitario

    return render(request, 'producto/detalle.html', {
        'producto': producto,
        'proveedores_asignados': proveedores_asignados,
        'proveedores_disponibles': proveedores_disponibles,
        'margen': margen,
    })


def crear_producto_view(request):
    usuario, rol = get_usuario_session(request)
    if not usuario or rol != 'Administrador':
        return redirect('login')

    proveedores = Proveedor.objects.filter(es_activo=True).order_by('nombre')
    categorias = Producto.objects.values_list('categoria', flat=True).distinct().exclude(categoria=None).order_by('categoria')

    if request.method == 'POST':
        nombre = request.POST.get('nombre')
        descripcion = request.POST.get('descripcion')
        costo_unitario = request.POST.get('costo_unitario')
        precio_venta = request.POST.get('precio_venta')
        stock = request.POST.get('stock')
        codigo_barras = request.POST.get('codigo_barras')
        fecha_vencimiento = request.POST.get('fecha_vencimiento') or None
        id_proveedor = request.POST.get('proveedor')
        categoria = request.POST.get('categoria_nueva') or request.POST.get('categoria')
        imagen = request.FILES.get('imagen')

        if not nombre or not costo_unitario:
            messages.error(request, 'Nombre y costo unitario son obligatorios.')
            return render(request, 'producto/crear.html', {
                'proveedores': proveedores,
                'categorias': categorias,
            })

        nuevo = Producto.objects.create(
            nombre=nombre,
            descripcion=descripcion,
            costo_unitario=costo_unitario,
            precio_venta=precio_venta or None,
            stock=stock or 0,
            categoria=categoria,
            codigo_barras=codigo_barras,
            fecha_vencimiento=fecha_vencimiento,
            estado='activo',
            es_activo=True,
            imagen=imagen,
            create_at=timezone.now(),
            update_at=timezone.now(),
        )

        if id_proveedor:
            proveedor = Proveedor.objects.get(id_proveedor=id_proveedor)
            ProveedorProducto.objects.create(
                proveedor=proveedor,
                producto=nuevo,
                precio_proveedor=costo_unitario,
            )

        messages.success(request, f'Producto {nombre} creado correctamente.')
        return redirect('lista_productos')

    return render(request, 'producto/crear.html', {
        'proveedores': proveedores,
        'categorias': categorias,
    })

def editar_producto_view(request, id):
    usuario, rol = get_usuario_session(request)
    if not usuario or rol != 'Administrador':
        return redirect('login')

    producto = get_object_or_404(Producto, id_producto=id)
    categorias = Producto.objects.values_list('categoria', flat=True).distinct().exclude(categoria=None).order_by('categoria')
    proveedores_asignados = ProveedorProducto.objects.filter(producto=producto).select_related('proveedor')
    proveedores_disponibles = Proveedor.objects.filter(es_activo=True).exclude(
        id_proveedor__in=[pp.proveedor.id_proveedor for pp in proveedores_asignados]
    )

    if request.method == 'POST':
        producto.nombre = request.POST.get('nombre')
        producto.descripcion = request.POST.get('descripcion')
        producto.costo_unitario = request.POST.get('costo_unitario')
        producto.precio_venta = request.POST.get('precio_venta') or None
        producto.stock = request.POST.get('stock') or 0
        producto.categoria = request.POST.get('categoria_nueva') or request.POST.get('categoria')
        producto.codigo_barras = request.POST.get('codigo_barras')
        producto.fecha_vencimiento = request.POST.get('fecha_vencimiento') or None
        producto.precio_promocion = request.POST.get('precio_promocion') or None
        producto.fecha_inicio_promo = request.POST.get('fecha_inicio_promo') or None
        producto.fecha_fin_promo = request.POST.get('fecha_fin_promo') or None
        producto.update_at = timezone.now()

        if request.FILES.get('imagen'):
            producto.imagen = request.FILES.get('imagen')

        producto.save()
        messages.success(request, f'Producto {producto.nombre} actualizado correctamente.')
        return redirect('detalle_producto', id=id)

    return render(request, 'producto/editar.html', {
        'producto': producto,
        'categorias': categorias,
        'proveedores_asignados': proveedores_asignados,
        'proveedores_disponibles': proveedores_disponibles,
    })

def desactivar_producto_view(request, id):
    usuario, rol = get_usuario_session(request)
    if not usuario or rol != 'Administrador':
        return redirect('login')

    producto = get_object_or_404(Producto, id_producto=id)

    # Si está inactivo lo reactiva directamente
    if not producto.es_activo:
        producto.es_activo = True
        producto.estado = 'activo'
        producto.motivo_desactivacion = None
        producto.save(update_fields=['es_activo', 'estado', 'motivo_desactivacion'])
        messages.success(request, f'Producto {producto.nombre} reactivado correctamente.')
        return redirect('lista_productos')

    if request.method == 'POST':
        motivo = request.POST.get('motivo')
        estado_nuevo = request.POST.get('estado')
        producto.es_activo = False
        producto.estado = estado_nuevo or 'inactivo'
        producto.motivo_desactivacion = motivo
        producto.save(update_fields=['es_activo', 'estado', 'motivo_desactivacion'])
        messages.success(request, f'Producto {producto.nombre} desactivado.')
        return redirect('lista_productos')

    return render(request, 'producto/desactivar.html', {
        'producto': producto,
        'motivos': [('bajo_stock', 'Bajo stock'), ('vencido', 'Producto vencido'), ('otro', 'Otro')],
        'estados': [('inactivo', 'Inactivo'), ('agotado', 'Agotado'), ('vencido', 'Vencido')],
    })


def eliminar_producto_view(request, id):
    usuario, rol = get_usuario_session(request)
    if not usuario or rol != 'Administrador':
        return redirect('login')

    producto = get_object_or_404(Producto, id_producto=id)
    nombre = producto.nombre
    producto.delete()
    messages.success(request, f'Producto {nombre} eliminado correctamente.')
    return redirect('lista_productos')


def asignar_proveedor_view(request, id):
    usuario, rol = get_usuario_session(request)
    if not usuario or rol != 'Administrador':
        return redirect('login')

    producto = get_object_or_404(Producto, id_producto=id)

    if request.method == 'POST':
        id_proveedor = request.POST.get('proveedor')
        precio_proveedor = request.POST.get('precio_proveedor') or None
        proveedor = get_object_or_404(Proveedor, id_proveedor=id_proveedor)

        ProveedorProducto.objects.get_or_create(
            proveedor=proveedor,
            producto=producto,
            defaults={'precio_proveedor': precio_proveedor}
        )
        messages.success(request, f'Proveedor {proveedor.nombre} asignado correctamente.')
        return redirect('detalle_producto', id=id)

    return redirect('detalle_producto', id=id)


def exportar_excel_productos(request):
    usuario, rol = get_usuario_session(request)
    if not usuario or rol != 'Administrador':
        return redirect('login')

    productos = Producto.objects.all().order_by('nombre')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Productos'

    headers = ['ID', 'Nombre', 'Categoría', 'Costo', 'Precio Venta', 'Stock', 'Estado', 'En Promoción']
    header_fill = PatternFill(start_color='28A745', end_color='28A745', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for row, p in enumerate(productos, 2):
        ws.cell(row=row, column=1, value=p.id_producto)
        ws.cell(row=row, column=2, value=p.nombre)
        ws.cell(row=row, column=3, value=p.categoria or '')
        ws.cell(row=row, column=4, value=float(p.costo_unitario))
        ws.cell(row=row, column=5, value=float(p.precio_venta) if p.precio_venta else '')
        ws.cell(row=row, column=6, value=p.stock or 0)
        ws.cell(row=row, column=7, value=p.estado)
        ws.cell(row=row, column=8, value='Sí' if p.en_promocion else 'No')

    for col in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 4

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="productos.xlsx"'
    wb.save(response)
    return response


def exportar_pdf_productos(request):
    usuario, rol = get_usuario_session(request)
    if not usuario or rol != 'Administrador':
        return redirect('login')

    # Filtros
    filtro_categoria = request.GET.get('categoria', '')
    filtro_estado = request.GET.get('estado', '')
    filtro_stock = request.GET.get('stock', '')

    productos = Producto.objects.all().order_by('nombre')
    if filtro_categoria:
        productos = productos.filter(categoria=filtro_categoria)
    if filtro_estado:
        productos = productos.filter(estado=filtro_estado)
    if filtro_stock == 'bajo':
        productos = [p for p in productos if p.stock_bajo]

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="productos.pdf"'

    doc = SimpleDocTemplate(response, pagesize=landscape(A4), leftMargin=30, rightMargin=30, topMargin=30, bottomMargin=30)
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle('Title', parent=styles['Title'], fontSize=18,
        textColor=colors.HexColor('#28A745'), spaceAfter=4, fontName='Helvetica-Bold', alignment=1)
    subtitle_style = ParagraphStyle('Subtitle', parent=styles['Normal'], fontSize=9,
        textColor=colors.HexColor('#666666'), alignment=1, spaceAfter=2)

    elements = []

    logo_path = os.path.join(settings.BASE_DIR, 'gestion', 'static', 'gestion', 'img', 'logo.jpg')
    if os.path.exists(logo_path):
        from reportlab.platypus import Image as RLImage
        logo = RLImage(logo_path, width=80, height=80)
        logo2 = RLImage(logo_path, width=80, height=80)
        header_data = [[logo, Paragraph('Reporte de Productos', title_style), logo2]]
        header_table = Table(header_data, colWidths=[90, 580, 90])
        header_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ALIGN', (0, 0), (0, 0), 'LEFT'),
            ('ALIGN', (1, 0), (1, 0), 'CENTER'),
            ('ALIGN', (2, 0), (2, 0), 'RIGHT'),
        ]))
        elements.append(header_table)
    else:
        elements.append(Paragraph('Reporte de Productos', title_style))

    # Subtítulo con filtros aplicados
    filtros_aplicados = []
    if filtro_categoria:
        filtros_aplicados.append(f'Categoría: {filtro_categoria}')
    if filtro_estado:
        filtros_aplicados.append(f'Estado: {filtro_estado}')
    if filtro_stock == 'bajo':
        filtros_aplicados.append('Stock: Bajo')

    filtros_texto = ' | '.join(filtros_aplicados) if filtros_aplicados else 'Sin filtros aplicados'

    elements.append(Paragraph('SenaFOOD - Sistema de Gestión', subtitle_style))
    elements.append(Paragraph(f'Generado el {datetime.now().strftime("%d/%m/%Y a las %H:%M")} por {request.session.get("usuario_nombre", "Administrador")}', subtitle_style))
    elements.append(Paragraph(f'Filtros: {filtros_texto}', subtitle_style))
    elements.append(Spacer(1, 6))
    elements.append(HRFlowable(width="100%", thickness=2, color=colors.HexColor('#28A745')))
    elements.append(Spacer(1, 8))

    total = len(productos) if isinstance(productos, list) else productos.count()
    activos = sum(1 for p in productos if p.es_activo)
    inactivos = sum(1 for p in productos if not p.es_activo)

    stat_style = ParagraphStyle('Stat', parent=styles['Normal'], fontSize=9, textColor=colors.HexColor('#333333'), alignment=1)
    stats_data = [[
        Paragraph(f'<b>Total Productos:</b> {total}', stat_style),
        Paragraph(f'<b>Activos:</b> {activos}', stat_style),
        Paragraph(f'<b>Inactivos:</b> {inactivos}', stat_style),
    ]]
    stats_table = Table(stats_data, colWidths=[250, 150, 150])
    stats_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#f0fff4')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#a8d5b5')),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
    ]))
    elements.append(stats_table)
    elements.append(Spacer(1, 12))

    data = [['ID', 'Nombre', 'Categoría', 'Costo', 'Precio Venta', 'Stock', 'Estado', 'Promoción']]
    for p in productos:
        data.append([
            str(p.id_producto),
            p.nombre,
            p.categoria or '',
            f'${p.costo_unitario:,.0f}',
            f'${p.precio_venta:,.0f}' if p.precio_venta else '—',
            str(p.stock_actual),
            p.estado,
            'Sí' if p.en_promocion else 'No',
        ])

    table = Table(data, repeatRows=1, colWidths=[25, 120, 90, 70, 80, 40, 60, 60])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#28A745')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0fff4')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#a8d5b5')),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(table)

    def add_footer(canvas, doc):
        canvas.saveState()
        canvas.setFont('Helvetica', 8)
        canvas.setFillColor(colors.HexColor('#888888'))
        canvas.drawString(30, 15, 'SenaFOOD - Reporte de Productos')
        canvas.drawRightString(landscape(A4)[0] - 30, 15, f'Página {doc.page}')
        canvas.setStrokeColor(colors.HexColor('#28A745'))
        canvas.setLineWidth(1)
        canvas.line(30, 25, landscape(A4)[0] - 30, 25)
        canvas.restoreState()

    doc.build(elements, onFirstPage=add_footer, onLaterPages=add_footer)
    return response