from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.utils import timezone
from django.utils.timezone import localtime
from django.db.models import Sum, Count, Avg, F
from datetime import date, timedelta, datetime
from gestion.models import Carrito, Detallecarrito, Usuario

import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import landscape, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from django.conf import settings
import os
import pytz


def sesion_requerida_admin(view_func):
    from functools import wraps
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if 'usuario_id' not in request.session:
            return redirect('login')
        if request.session.get('usuario_rol') != 'Administrador':
            return redirect('dashboard')
        return view_func(request, *args, **kwargs)
    return wrapper


# ─────────────────────────────────────────
# LISTA VENTAS
# ─────────────────────────────────────────
@sesion_requerida_admin
def lista_ventas(request):
    from django.db.models import Q

    filtro_fecha_inicio = request.GET.get('fecha_inicio', '')
    filtro_fecha_fin    = request.GET.get('fecha_fin', '')
    filtro_cliente      = request.GET.get('cliente', '')

    ahora         = datetime.utcnow() - timedelta(hours=5)
    hoy           = ahora.date()
    semana        = hoy - timedelta(days=7)
    mes           = hoy.replace(day=1)

    hoy_inicio    = datetime.combine(hoy,    datetime.min.time())
    hoy_fin       = datetime.combine(hoy,    datetime.max.time())
    semana_inicio = datetime.combine(semana, datetime.min.time())
    mes_inicio    = datetime.combine(mes,    datetime.min.time())

    # ── Totales ──
    total_hoy = Carrito.objects.filter(estado='entregado').filter(
        Q(fecha_entrega__gte=hoy_inicio, fecha_entrega__lte=hoy_fin) |
        Q(fecha_entrega__isnull=True, update_at__gte=hoy_inicio, update_at__lte=hoy_fin)
    ).aggregate(t=Sum('total'))['t'] or 0

    total_semana = Carrito.objects.filter(estado='entregado').filter(
        Q(fecha_entrega__gte=semana_inicio) |
        Q(fecha_entrega__isnull=True, update_at__gte=semana_inicio)
    ).aggregate(t=Sum('total'))['t'] or 0

    total_mes = Carrito.objects.filter(estado='entregado').filter(
        Q(fecha_entrega__gte=mes_inicio) |
        Q(fecha_entrega__isnull=True, update_at__gte=mes_inicio)
    ).aggregate(t=Sum('total'))['t'] or 0

    total_pedidos = Carrito.objects.filter(estado='entregado').count()

    # ── Lista ventas ──
    ventas = Carrito.objects.filter(
        estado='entregado'
    ).select_related('usuario').order_by('-update_at')

    if filtro_fecha_inicio:
        fecha_inicio_dt = datetime.combine(
            date.fromisoformat(filtro_fecha_inicio),
            datetime.min.time()
        )
        ventas = ventas.filter(
            Q(fecha_entrega__gte=fecha_inicio_dt) |
            Q(fecha_entrega__isnull=True, update_at__gte=fecha_inicio_dt)
        )
    if filtro_fecha_fin:
        fecha_fin_dt = datetime.combine(
            date.fromisoformat(filtro_fecha_fin),
            datetime.max.time()
        )
        ventas = ventas.filter(
            Q(fecha_entrega__lte=fecha_fin_dt) |
            Q(fecha_entrega__isnull=True, update_at__lte=fecha_fin_dt)
        )
    if filtro_cliente:
        ventas = ventas.filter(
            Q(usuario__nombre__icontains=filtro_cliente) |
            Q(usuario__apellido__icontains=filtro_cliente)
        )

    # ── Producto más vendido ──
    producto_top = Detallecarrito.objects.filter(
        id_carrito__estado='entregado'
    ).values('id_producto__nombre').annotate(
        total_vendido=Sum('cantidad')
    ).order_by('-total_vendido').first()

    # ── Tiempo promedio ──
    carritos_con_fechas = Carrito.objects.filter(
        estado='entregado',
        fecha_confirmacion__isnull=False,
        fecha_entrega__isnull=False
    )
    tiempo_promedio = None
    if carritos_con_fechas.exists():
        tiempos = [
            (c.fecha_entrega - c.fecha_confirmacion).total_seconds() / 60
            for c in carritos_con_fechas
        ]
        tiempo_promedio = round(sum(tiempos) / len(tiempos), 1)

    # ── Gráfica días de la semana (últimos 7 días) ──
    dias_semana    = ['Lun', 'Mar', 'Mié', 'Jue', 'Vie', 'Sáb', 'Dom']
    ventas_por_dia = [0] * 7

    hace_7_dias = datetime.combine(semana, datetime.min.time())

    for c in Carrito.objects.filter(estado='entregado').filter(
        Q(fecha_entrega__gte=hace_7_dias) |
        Q(fecha_entrega__isnull=True, update_at__gte=hace_7_dias)
    ):
        fecha = c.fecha_entrega or c.update_at
        if fecha:
            dia = fecha.weekday()
            ventas_por_dia[dia] += float(c.total or 0)

    # ── Gráfica categorías ──
    ventas_categoria = Detallecarrito.objects.filter(
        id_carrito__estado='entregado'
    ).values('id_producto__categoria').annotate(
        total=Sum(F('cantidad') * F('precio_unitario'))
    ).order_by('-total')[:6]

    categorias_labels  = [v['id_producto__categoria'] or 'Sin categoría' for v in ventas_categoria]
    categorias_valores = [float(v['total'] or 0) for v in ventas_categoria]

    # ── Top productos ──
    top_productos = Detallecarrito.objects.filter(
        id_carrito__estado='entregado'
    ).values('id_producto__nombre').annotate(
        cantidad=Sum('cantidad')
    ).order_by('-cantidad')[:5]

    top_nombres    = [p['id_producto__nombre'] for p in top_productos]
    top_cantidades = [p['cantidad'] for p in top_productos]

    for v in ventas:
        if v.fecha_confirmacion and v.fecha_entrega:
            v.tiempo_entrega_min = round(
                (v.fecha_entrega - v.fecha_confirmacion).total_seconds() / 60, 1
            )
        else:
            v.tiempo_entrega_min = None

    context = {
        'ventas':              ventas,
        'total_hoy':           total_hoy,
        'total_semana':        total_semana,
        'total_mes':           total_mes,
        'total_pedidos':       total_pedidos,
        'producto_top':        producto_top,
        'tiempo_promedio':     tiempo_promedio,
        'filtro_fecha_inicio': filtro_fecha_inicio,
        'filtro_fecha_fin':    filtro_fecha_fin,
        'filtro_cliente':      filtro_cliente,
        'dias_semana':         dias_semana,
        'ventas_por_dia':      ventas_por_dia,
        'categorias_labels':   categorias_labels,
        'categorias_valores':  categorias_valores,
        'top_nombres':         top_nombres,
        'top_cantidades':      top_cantidades,
        'nombre_usuario':      request.session.get('usuario_nombre', ''),
        'rol_usuario':         request.session.get('usuario_rol', ''),
    }
    return render(request, 'ventas/lista.html', context)


# ─────────────────────────────────────────
# DETALLE VENTA
# ─────────────────────────────────────────
@sesion_requerida_admin
def detalle_venta(request, id_carrito):
    carrito  = get_object_or_404(Carrito, pk=id_carrito, estado='entregado')
    detalles = Detallecarrito.objects.filter(
        id_carrito=carrito
    ).select_related('id_producto')

    tiempo_entrega = None
    if carrito.fecha_confirmacion and carrito.fecha_entrega:
        diff = (carrito.fecha_entrega - carrito.fecha_confirmacion).total_seconds() / 60
        tiempo_entrega = round(diff, 1)

    return render(request, 'ventas/detalle.html', {
        'carrito':        carrito,
        'detalles':       detalles,
        'tiempo_entrega': tiempo_entrega,
        'nombre_usuario': request.session.get('usuario_nombre', ''),
        'rol_usuario':    request.session.get('usuario_rol', ''),
    })


# ─────────────────────────────────────────
# EXPORTAR PDF
# ─────────────────────────────────────────
@sesion_requerida_admin
def exportar_pdf_ventas(request):
    from django.db.models import Q

    filtro_fecha_inicio = request.GET.get('fecha_inicio', '')
    filtro_fecha_fin    = request.GET.get('fecha_fin', '')

    ventas = Carrito.objects.filter(
        estado='entregado'
    ).select_related('usuario').order_by('-update_at')

    if filtro_fecha_inicio:
        ventas = ventas.filter(
            Q(fecha_entrega__date__gte=filtro_fecha_inicio) |
            Q(fecha_entrega__isnull=True, update_at__date__gte=filtro_fecha_inicio)
        )
    if filtro_fecha_fin:
        ventas = ventas.filter(
            Q(fecha_entrega__date__lte=filtro_fecha_fin) |
            Q(fecha_entrega__isnull=True, update_at__date__lte=filtro_fecha_fin)
        )

    buffer = io.BytesIO()
    doc    = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            leftMargin=30, rightMargin=30,
                            topMargin=30, bottomMargin=30)
    styles   = getSampleStyleSheet()
    elements = []

    title_style = ParagraphStyle('T', parent=styles['Title'], fontSize=16,
                                textColor=colors.HexColor('#28A745'),
                                fontName='Helvetica-Bold', alignment=1)
    sub_style   = ParagraphStyle('S', parent=styles['Normal'], fontSize=9,
                                textColor=colors.HexColor('#666'), alignment=1)

    logo_path = os.path.join(settings.BASE_DIR, 'gestion', 'static', 'gestion', 'img', 'logo.png')
    if os.path.exists(logo_path):
        from reportlab.platypus import Image as RLImage
        logo  = RLImage(logo_path, width=70, height=70)
        logo2 = RLImage(logo_path, width=70, height=70)
        ht    = Table([[logo, Paragraph('Reporte de Ventas', title_style), logo2]],
                    colWidths=[80, 600, 80])
        ht.setStyle(TableStyle([('VALIGN',(0,0),(-1,-1),'MIDDLE'),
                                ('ALIGN',(1,0),(1,0),'CENTER')]))
        elements.append(ht)
    else:
        elements.append(Paragraph('Reporte de Ventas', title_style))

    # Rango de fechas
    if filtro_fecha_inicio and filtro_fecha_fin:
        rango = f'Período: {filtro_fecha_inicio} al {filtro_fecha_fin}'
    elif filtro_fecha_inicio:
        rango = f'Desde: {filtro_fecha_inicio}'
    elif filtro_fecha_fin:
        rango = f'Hasta: {filtro_fecha_fin}'
    else:
        rango = 'Período: Todos los registros'

    elements.append(Paragraph(rango, sub_style))
    elements.append(Paragraph(f'Generado el {datetime.now().strftime("%d/%m/%Y %H:%M")}', sub_style))
    elements.append(Spacer(1, 8))

    from reportlab.platypus import HRFlowable
    elements.append(HRFlowable(width="100%", thickness=2, color=colors.HexColor('#28A745')))
    elements.append(Spacer(1, 8))

    total_general = sum(float(v.total or 0) for v in ventas)
    stat_style = ParagraphStyle('St', parent=styles['Normal'], fontSize=9,
                                textColor=colors.HexColor('#333'), alignment=1)
    stats = Table([[
        Paragraph(f'<b>Total ventas:</b> {ventas.count()}', stat_style),
        Paragraph(f'<b>Total ingresos:</b> ${total_general:,.0f}', stat_style),
    ]], colWidths=[370, 370])
    stats.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,-1), colors.HexColor('#f0fff4')),
        ('GRID',(0,0),(-1,-1), 0.5, colors.HexColor('#a8d5b5')),
        ('TOPPADDING',(0,0),(-1,-1), 8),
        ('BOTTOMPADDING',(0,0),(-1,-1), 8),
        ('ALIGN',(0,0),(-1,-1),'CENTER'),
    ]))
    elements.append(stats)
    elements.append(Spacer(1, 12))

    data = [['#', 'Pedido', 'Cliente', 'Total', 'Método pago', 'Tiempo entrega']]
    for i, v in enumerate(ventas, 1):
        fecha = v.fecha_entrega.strftime('%d/%m/%Y %H:%M') if v.fecha_entrega else v.update_at.strftime('%d/%m/%Y %H:%M') if v.update_at else '—'

        if v.fecha_confirmacion and v.fecha_entrega:
            minutos = round((v.fecha_entrega - v.fecha_confirmacion).total_seconds() / 60, 1)
            tiempo  = f'{minutos} min'
        else:
            tiempo = '—'

        data.append([
            str(i),
            f'#{v.id_carrito}',
            f'{v.usuario.nombre} {v.usuario.apellido}',
            f'${float(v.total or 0):,.0f}',
            v.metodopago or '—',
            tiempo,
        ])
    tabla = Table(data, repeatRows=1, colWidths=[30, 60, 180, 120, 100, 100])
    tabla.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,0), colors.HexColor('#28A745')),
        ('TEXTCOLOR',(0,0),(-1,0), colors.white),
        ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
        ('FONTSIZE',(0,0),(-1,0), 9),
        ('FONTSIZE',(0,1),(-1,-1), 8),
        ('ALIGN',(0,0),(-1,-1),'CENTER'),
        ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
        ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white, colors.HexColor('#f0fff4')]),
        ('GRID',(0,0),(-1,-1), 0.5, colors.HexColor('#a8d5b5')),
        ('TOPPADDING',(0,0),(-1,-1), 5),
        ('BOTTOMPADDING',(0,0),(-1,-1), 5),
    ]))
    elements.append(tabla)

    def footer(canvas, doc):
        canvas.saveState()
        canvas.setFont('Helvetica', 8)
        canvas.setFillColor(colors.HexColor('#888'))
        canvas.drawString(30, 15, 'SenaFOOD - Reporte de Ventas')
        canvas.drawRightString(landscape(A4)[0]-30, 15, f'Página {doc.page}')
        canvas.setStrokeColor(colors.HexColor('#28A745'))
        canvas.setLineWidth(1)
        canvas.line(30, 25, landscape(A4)[0]-30, 25)
        canvas.restoreState()

    doc.build(elements, onFirstPage=footer, onLaterPages=footer)
    buffer.seek(0)
    return HttpResponse(buffer, content_type='application/pdf',
        headers={'Content-Disposition': 'attachment; filename="ventas.pdf"'})


# ─────────────────────────────────────────
# EXPORTAR EXCEL
# ─────────────────────────────────────────
@sesion_requerida_admin
def exportar_excel_ventas(request):
    filtro_fecha_inicio = request.GET.get('fecha_inicio', '')
    filtro_fecha_fin    = request.GET.get('fecha_fin', '')

    ventas = Carrito.objects.filter(estado='entregado').select_related('usuario').order_by('-fecha_entrega')
    if filtro_fecha_inicio:
        ventas = ventas.filter(fecha_entrega__date__gte=filtro_fecha_inicio)
    if filtro_fecha_fin:
        ventas = ventas.filter(fecha_entrega__date__lte=filtro_fecha_fin)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Ventas'

    verde_fill  = PatternFill('solid', fgColor='28a745')
    blanco_font = Font(color='FFFFFF', bold=True)
    centro      = Alignment(horizontal='center', vertical='center')

    ws['A1'] = 'Reporte de Ventas — SenaFOOD'
    ws['A1'].font = Font(bold=True, size=13, color='28a745')
    ws.append([])

    encabezados = ['#', 'Pedido', 'Cliente', 'Fecha entrega', 'Total', 'Método pago', 'Tiempo entrega (min)']
    ws.append(encabezados)

    for cell in ws[3]:
        cell.fill      = verde_fill
        cell.font      = blanco_font
        cell.alignment = centro

    for i, v in enumerate(ventas, 1):
        fecha = localtime(v.fecha_entrega).strftime('%d/%m/%Y %H:%M') if v.fecha_entrega else '—'
        tiempo = ''
        if v.fecha_confirmacion and v.fecha_entrega:
            tiempo = round((v.fecha_entrega - v.fecha_confirmacion).total_seconds() / 60, 1)
        ws.append([
            i,
            f'#{v.id_carrito}',
            f'{v.usuario.nombre} {v.usuario.apellido}',
            fecha,
            float(v.total or 0),
            v.metodopago or '—',
            tiempo,
        ])

    anchos = [5, 10, 30, 20, 15, 15, 20]
    for col, ancho in enumerate(anchos, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = ancho

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return HttpResponse(buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename="ventas.xlsx"'})