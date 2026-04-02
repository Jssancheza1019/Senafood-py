import openpyxl
import os
from django.conf import settings
from reportlab.lib.styles import ParagraphStyle
from openpyxl.styles import Font, PatternFill, Alignment
from django.http import HttpResponse
from reportlab.lib.pagesizes import landscape, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from datetime import datetime
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from .models import Proveedor
from gestion.models import Usuario


def get_usuario_session(request):
    if 'usuario_id' not in request.session:
        return None, None
    try:
        usuario = Usuario.objects.get(id_usuario=request.session['usuario_id'])
        rol = usuario.rol.nombre_rol if usuario.rol else None
        return usuario, rol
    except Usuario.DoesNotExist:
        return None, None


def lista_proveedores_view(request):
    usuario, rol = get_usuario_session(request)
    if not usuario or rol != 'Administrador':
        return redirect('login')

    filtro_estado = request.GET.get('estado', '')
    proveedores = Proveedor.objects.all().order_by('-create_at')

    if filtro_estado == '1':
        proveedores = proveedores.filter(es_activo=True)
    elif filtro_estado == '0':
        proveedores = proveedores.filter(es_activo=False)

    # Conteos cards
    total_activos   = Proveedor.objects.filter(es_activo=True).count()
    total_inactivos = Proveedor.objects.filter(es_activo=False).count()

    return render(request, 'proveedor/lista.html', {
        'proveedores':      proveedores,
        'filtro_estado':    filtro_estado,
        'total_activos':    total_activos,
        'total_inactivos':  total_inactivos,
    })


def crear_proveedor_view(request):
    usuario, rol = get_usuario_session(request)
    if not usuario or rol != 'Administrador':
        return redirect('login')

    if request.method == 'POST':
        nombre = request.POST.get('nombre')
        nit = request.POST.get('nit')
        email = request.POST.get('email')
        telefono = request.POST.get('telefono')
        direccion = request.POST.get('direccion')

        if not nombre or not nit:
            messages.error(request, 'Nombre y NIT son obligatorios.')
            return render(request, 'proveedor/crear.html')

        if Proveedor.objects.filter(nit=nit).exists():
            messages.error(request, f'El NIT "{nit}" ya está registrado.')
            return render(request, 'proveedor/crear.html')

        Proveedor.objects.create(
            nombre=nombre,
            nit=nit,
            email=email,
            telefono=telefono,
            direccion=direccion,
            es_activo=True,
        )
        messages.success(request, f'Proveedor {nombre} creado correctamente.')
        return redirect('lista_proveedores')

    return render(request, 'proveedor/crear.html')


def editar_proveedor_view(request, id):
    usuario, rol = get_usuario_session(request)
    if not usuario or rol != 'Administrador':
        return redirect('login')

    proveedor = get_object_or_404(Proveedor, id_proveedor=id)

    if request.method == 'POST':
        proveedor.nombre = request.POST.get('nombre')
        proveedor.nit = request.POST.get('nit')
        proveedor.email = request.POST.get('email')
        proveedor.telefono = request.POST.get('telefono')
        proveedor.direccion = request.POST.get('direccion')
        proveedor.es_activo = request.POST.get('es_activo') == '1'
        proveedor.save()
        messages.success(request, f'Proveedor {proveedor.nombre} actualizado correctamente.')
        return redirect('lista_proveedores')

    return render(request, 'proveedor/editar.html', {
        'proveedor': proveedor,
    })


def eliminar_proveedor_view(request, id):
    usuario, rol = get_usuario_session(request)
    if not usuario or rol != 'Administrador':
        return redirect('login')

    proveedor = get_object_or_404(Proveedor, id_proveedor=id)
    
    if proveedor.es_activo:
        proveedor.es_activo = False
        messages.success(request, f'Proveedor {proveedor.nombre} desactivado correctamente.')
    else:
        proveedor.es_activo = True
        messages.success(request, f'Proveedor {proveedor.nombre} reactivado correctamente.')
    
    proveedor.save()
    return redirect('lista_proveedores')

def detalle_proveedor_view(request, id):
    usuario, rol = get_usuario_session(request)
    if not usuario or rol != 'Administrador':
        return redirect('login')

    proveedor = get_object_or_404(Proveedor, id_proveedor=id)

    return render(request, 'proveedor/detalle.html', {
        'proveedor': proveedor,
    })

def exportar_excel_proveedores(request):
    usuario, rol = get_usuario_session(request)
    if not usuario or rol != 'Administrador':
        return redirect('login')

    proveedores = Proveedor.objects.all().order_by('nombre')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Proveedores'

    headers = ['ID', 'Nombre', 'NIT', 'Email', 'Teléfono', 'Dirección', 'Estado']
    header_fill = PatternFill(start_color='28A745', end_color='28A745', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for row, p in enumerate(proveedores, 2):
        ws.cell(row=row, column=1, value=p.id_proveedor)
        ws.cell(row=row, column=2, value=p.nombre)
        ws.cell(row=row, column=3, value=p.nit)
        ws.cell(row=row, column=4, value=p.email or '')
        ws.cell(row=row, column=5, value=p.telefono or '')
        ws.cell(row=row, column=6, value=p.direccion or '')
        ws.cell(row=row, column=7, value='Activo' if p.es_activo else 'Inactivo')

    for col in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 4

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="proveedores.xlsx"'
    wb.save(response)
    return response


def exportar_pdf_proveedores(request):
    usuario, rol = get_usuario_session(request)
    if not usuario or rol != 'Administrador':
        return redirect('login')

    proveedores = Proveedor.objects.all().order_by('nombre')

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="proveedores.pdf"'

    doc = SimpleDocTemplate(
        response,
        pagesize=landscape(A4),
        leftMargin=30, rightMargin=30,
        topMargin=30, bottomMargin=30
    )

    styles = getSampleStyleSheet()
    elements = []

    # Estilo título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Title'],
        fontSize=18,
        textColor=colors.HexColor('#28A745'),
        spaceAfter=4,
        fontName='Helvetica-Bold',
        alignment=1,
    )
    subtitle_style = ParagraphStyle(
        'Subtitle',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.HexColor('#666666'),
        alignment=1,
        spaceAfter=2,
    )
    stat_style = ParagraphStyle(
        'Stat',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.HexColor('#333333'),
        alignment=1,
    )

    # Header con logo
    logo_path = os.path.join(settings.BASE_DIR, 'gestion', 'static', 'gestion', 'img', 'logo.png')

    if os.path.exists(logo_path):
        from reportlab.platypus import Image as RLImage
        logo = RLImage(logo_path, width=80, height=80)
        logo2 = RLImage(logo_path, width=80, height=80)
        header_data = [[logo, Paragraph('Reporte de Proveedores', title_style), logo2]]
        header_table = Table(header_data, colWidths=[90, 580, 90])
        header_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ALIGN', (0, 0), (0, 0), 'LEFT'),
            ('ALIGN', (1, 0), (1, 0), 'CENTER'),
            ('ALIGN', (2, 0), (2, 0), 'RIGHT'),
        ]))
        elements.append(header_table)
    else:
        elements.append(Paragraph('Reporte de Proveedores', title_style))

    elements.append(Paragraph('SenaFOOD - Sistema de Gestión', subtitle_style))
    elements.append(Paragraph(f'Generado el {datetime.now().strftime("%d/%m/%Y a las %H:%M")} por {request.session.get("usuario_nombre", "Administrador")}', subtitle_style))
    elements.append(Spacer(1, 6))

    # Línea verde decorativa
    from reportlab.platypus import HRFlowable
    elements.append(HRFlowable(width="100%", thickness=2, color=colors.HexColor('#28A745')))
    elements.append(Spacer(1, 8))

    # Estadísticas
    total = proveedores.count()
    activos = proveedores.filter(es_activo=True).count()
    inactivos = proveedores.filter(es_activo=False).count()

    stats_data = [[
        Paragraph(f'<b>Total Proveedores:</b> {total}', stat_style),
        Paragraph(f'<b>Activos:</b> {activos}', stat_style),
        Paragraph(f'<b>Inactivos:</b> {inactivos}', stat_style),
    ]]
    stats_table = Table(stats_data, colWidths=[250, 150, 150])
    stats_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#f0fff4')),
        ('BORDER', (0, 0), (-1, -1), 0.5, colors.HexColor('#a8d5b5')),
        ('ROUNDEDCORNERS', [5]),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
    ]))
    elements.append(stats_table)
    elements.append(Spacer(1, 12))

    # Tabla de datos
    data = [['ID', 'Nombre', 'NIT', 'Email', 'Teléfono', 'Dirección', 'Estado']]
    for p in proveedores:
        data.append([
            str(p.id_proveedor),
            p.nombre or '',
            p.nit or '',
            p.email or '',
            p.telefono or '',
            p.direccion or '',
            'Activo' if p.es_activo else 'Inactivo',
        ])

    table = Table(data, repeatRows=1, colWidths=[30, 100, 80, 130, 80, 150, 50])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#28A745')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0fff4')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#a8d5b5')),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(table)

    # Footer con número de página
    def add_footer(canvas, doc):
        canvas.saveState()
        canvas.setFont('Helvetica', 8)
        canvas.setFillColor(colors.HexColor('#888888'))
        canvas.drawString(30, 15, f'SenaFOOD - Reporte de Proveedores')
        canvas.drawRightString(landscape(A4)[0] - 30, 15, f'Página {doc.page}')
        canvas.setStrokeColor(colors.HexColor('#28A745'))
        canvas.setLineWidth(1)
        canvas.line(30, 25, landscape(A4)[0] - 30, 25)
        canvas.restoreState()

    doc.build(elements, onFirstPage=add_footer, onLaterPages=add_footer)
    return response