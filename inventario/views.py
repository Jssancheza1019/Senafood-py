from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.http import HttpResponse
from django.utils import timezone
from datetime import date
from functools import wraps

from .models import Inventario, InventarioDiario, DetalleInventarioDiario

import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import landscape, A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle


def sesion_requerida(view_func):
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if 'usuario_id' not in request.session:
            return redirect('login')
        return view_func(request, *args, **kwargs)
    return wrapper

# ─────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────



def recalcular_estado(producto, alerta_minimos=5):
    """
    Recalcula el estado del producto según stock y vencimiento.
    Retorna el string del nuevo estado.
    """
    hoy = date.today()
    if producto.fecha_vencimiento and producto.fecha_vencimiento < hoy:
        return 'vencido'
    if not producto.es_activo:
        return 'inactivo'
    stock = producto.stock or 0
    if stock <= 0:
        return 'agotado'
    if stock <= alerta_minimos:
        return 'agotado'
    return 'activo'


def get_alerta_minimos(id_producto, default=5):
    """
    Busca el alerta_minimos en la tabla inventario para ese producto.
    Si no existe usa el default.
    """
    try:
        inv = Inventario.objects.get(idproducto=id_producto)
        return inv.alerta_minimos
    except Inventario.DoesNotExist:
        return default


def verificar_rol(request):
    rol = request.session.get('usuario_rol', '')
    return rol.lower() in ['administrador', 'vendedor']
# ─────────────────────────────────────────────────────
# 1. LISTA PRINCIPAL
# ─────────────────────────────────────────────────────

@sesion_requerida
@sesion_requerida
def lista_inventario(request):
    from gestion.models import Producto

    filtro_estado    = request.GET.get('estado', '')
    filtro_stock     = request.GET.get('stock', '')
    filtro_categoria = request.GET.get('categoria', '')
    filtro_vencer    = request.GET.get('vencer', '')

    productos   = Producto.objects.all().order_by('nombre')
    inventarios = {inv.idproducto: inv for inv in Inventario.objects.all()}

    hoy            = date.today()
    inventario_hoy = InventarioDiario.objects.filter(fecha=hoy).first()

    categorias = Producto.objects.values_list(
        'categoria', flat=True
    ).distinct().exclude(categoria__isnull=True).exclude(categoria='').order_by('categoria')

    productos_data   = []
    total_alerta     = 0
    total_por_vencer = 0
    total_ok         = 0

    for p in productos:
        inv         = inventarios.get(p.id_producto)
        alerta_min  = inv.alerta_minimos if inv else 5
        estado_calc = recalcular_estado(p, alerta_min)
        en_alerta   = (p.stock or 0) <= alerta_min
        por_vencer  = (
            p.fecha_vencimiento and
            p.fecha_vencimiento >= hoy and
            (p.fecha_vencimiento - hoy).days <= 7
        )

        if estado_calc in ['agotado', 'vencido'] or en_alerta:
            total_alerta += 1
        if por_vencer:
            total_por_vencer += 1
        if estado_calc == 'activo' and not en_alerta:
            total_ok += 1

        productos_data.append({
            'producto':    p,
            'inv':         inv,
            'alerta_min':  alerta_min,
            'estado_calc': estado_calc,
            'en_alerta':   en_alerta,
            'por_vencer':  por_vencer,
        })

    # Aplicar filtros
    if filtro_estado:
        productos_data = [i for i in productos_data if i['estado_calc'] == filtro_estado]
    if filtro_stock == 'bajo':
        productos_data = [i for i in productos_data if i['en_alerta']]
    if filtro_stock == 'cero':
        productos_data = [i for i in productos_data if (i['producto'].stock or 0) == 0]
    if filtro_categoria:
        productos_data = [i for i in productos_data if i['producto'].categoria == filtro_categoria]
    if filtro_vencer == 'si':
        productos_data = [i for i in productos_data if i['por_vencer']]

    context = {
        'productos_data':   productos_data,
        'inventario_hoy':   inventario_hoy,
        'total_productos':  Producto.objects.count(),
        'total_alerta':     total_alerta,
        'total_por_vencer': total_por_vencer,
        'total_ok':         total_ok,
        'hoy':              hoy,
        'categorias':       categorias,
        'filtro_estado':    filtro_estado,
        'filtro_stock':     filtro_stock,
        'filtro_categoria': filtro_categoria,
        'filtro_vencer':    filtro_vencer,
    }
    return render(request, 'inventario/lista.html', context)

# ─────────────────────────────────────────────────────
# 2. REALIZAR INVENTARIO DIARIO
# ─────────────────────────────────────────────────────

@sesion_requerida
def realizar_inventario(request):
    from gestion.models import Producto

    if not verificar_rol(request):
        messages.error(request, 'No tienes permisos para realizar inventarios.')
        return redirect('lista_inventario')

    productos = Producto.objects.all().order_by('nombre')
    inventarios = {inv.idproducto: inv for inv in Inventario.objects.all()}
    hoy         = date.today()

    if request.method == 'POST':
        observaciones = request.POST.get('observaciones', '').strip()

        try:
            nombre_responsable = request.session.get('usuario_nombre', 'Sin nombre')
            usuario_id         = request.session.get('usuario_id', 0)

            # Crear cabecera
            inv_diario = InventarioDiario.objects.create(
                fecha              = hoy,
                realizado_por      = usuario_id,
                nombre_responsable = nombre_responsable,
                observaciones      = observaciones,
                create_at          = timezone.now(),
            )

            total_productos   = 0
            total_diferencias = 0

            for p in productos:
                campo       = f'stock_{p.id_producto}'
                stock_nuevo = request.POST.get(campo)

                if stock_nuevo is None or stock_nuevo == '':
                    continue

                stock_nuevo    = int(stock_nuevo)
                stock_anterior = p.stock or 0
                diferencia     = stock_nuevo - stock_anterior
                alerta_min     = get_alerta_minimos(p.id_producto)
                estado_anterior = p.estado

                # Guardar detalle
                DetalleInventarioDiario.objects.create(
                    inventario_diario = inv_diario,
                    id_producto       = p.id_producto,
                    nombre_producto   = p.nombre,
                    categoria         = p.categoria or '',
                    stock_anterior    = stock_anterior,
                    stock_fisico      = stock_nuevo,
                    diferencia        = diferencia,
                    alerta_minimos    = alerta_min,
                    costo_unitario    = p.costo_unitario or 0,
                    estado_anterior   = estado_anterior,
                    estado_nuevo      = '',
                )

                # Actualizar stock en producto
                p.stock     = stock_nuevo
                p.update_at = timezone.now()

                # Recalcular estado
                estado_nuevo = recalcular_estado(p, alerta_min)
                p.estado     = estado_nuevo

                # Actualizar también la tabla inventario si existe
                inv_base = inventarios.get(p.id_producto)
                if inv_base:
                    inv_base.stocktotal  = stock_nuevo
                    inv_base.valor_total = stock_nuevo * float(p.costo_unitario or 0)
                    inv_base.update_at   = timezone.now()
                    inv_base.ultima_revision = hoy
                    inv_base.save()

                p.save()

                # Actualizar estado_nuevo en el detalle recién creado
                DetalleInventarioDiario.objects.filter(
                    inventario_diario=inv_diario,
                    id_producto=p.id_producto
                ).update(estado_nuevo=estado_nuevo)

                total_productos   += 1
                if diferencia != 0:
                    total_diferencias += 1

            # Actualizar totales en cabecera
            inv_diario.total_productos   = total_productos
            inv_diario.total_diferencias = total_diferencias
            inv_diario.save()

            messages.success(
                request,
                f'Inventario del {hoy} guardado correctamente. '
                f'{total_productos} productos procesados, '
                f'{total_diferencias} con diferencias.'
            )
            return redirect('detalle_inventario_diario', pk=inv_diario.pk)

        except Exception as e:
            messages.error(request, f'Error al guardar el inventario: {str(e)}')

    # Construir lista para el formulario
    productos_form = []
    for p in productos:
        inv        = inventarios.get(p.id_producto)
        alerta_min = inv.alerta_minimos if inv else 5
        productos_form.append({
            'producto':   p,
            'inv':        inv,
            'alerta_min': alerta_min,
            'en_alerta':  (p.stock or 0) <= alerta_min,
        })

    return render(request, 'inventario/realizar.html', {
        'productos_form': productos_form,
        'hoy':            hoy,
    })


# ─────────────────────────────────────────────────────
# 3. HISTORIAL DE INVENTARIOS
# ─────────────────────────────────────────────────────

@sesion_requerida
def historial_inventario(request):
    inventarios = InventarioDiario.objects.all().order_by('-fecha', '-create_at')
    return render(request, 'inventario/historial.html', {
        'inventarios': inventarios,
    })


# ─────────────────────────────────────────────────────
# 4. DETALLE DE UN INVENTARIO DIARIO
# ─────────────────────────────────────────────────────

@sesion_requerida
def detalle_inventario_diario(request, pk):
    inv_diario = get_object_or_404(InventarioDiario, pk=pk)
    detalles   = inv_diario.detalles.all().order_by('nombre_producto')

    con_diferencia  = detalles.filter(diferencia__lt=0).count()
    con_aumento     = detalles.filter(diferencia__gt=0).count()
    sin_diferencia  = detalles.filter(diferencia=0).count()

    return render(request, 'inventario/detalle.html', {
        'inv_diario':     inv_diario,
        'detalles':       detalles,
        'con_diferencia': con_diferencia,
        'con_aumento':    con_aumento,
        'sin_diferencia': sin_diferencia,
    })


# ─────────────────────────────────────────────────────
# 5. EXPORTAR PDF — DETALLE DE UN INVENTARIO
# ─────────────────────────────────────────────────────

@sesion_requerida
def exportar_pdf(request, pk):
    import os
    from django.conf import settings
    from reportlab.platypus import Image as RLImage, HRFlowable
    from datetime import datetime

    inv_diario = get_object_or_404(InventarioDiario, pk=pk)
    detalles   = inv_diario.detalles.all().order_by('nombre_producto')

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="inventario_{inv_diario.fecha}.pdf"'

    doc = SimpleDocTemplate(
        response,
        pagesize=landscape(A4),
        leftMargin=30, rightMargin=30,
        topMargin=30, bottomMargin=30
    )

    styles = getSampleStyleSheet()
    elements = []

    # ── Estilos ──
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Title'],
        fontSize=18,
        textColor=colors.HexColor('#28A745'),
        spaceAfter=4,
        fontName='Helvetica-Bold',
        alignment=1,
    )
    subtitle_style = ParagraphStyle(
        'Subtitle',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.HexColor('#666666'),
        alignment=1,
        spaceAfter=2,
    )
    stat_style = ParagraphStyle(
        'Stat',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.HexColor('#333333'),
        alignment=1,
    )

    # ── Header con logo ──
    logo_path = os.path.join(settings.BASE_DIR, 'gestion', 'static', 'gestion', 'img', 'logo.png')

    if os.path.exists(logo_path):
        logo  = RLImage(logo_path, width=80, height=80)
        logo2 = RLImage(logo_path, width=80, height=80)
        header_data = [[logo, Paragraph('Reporte de Inventario Diario', title_style), logo2]]
        header_table = Table(header_data, colWidths=[90, 580, 90])
        header_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ALIGN', (0, 0), (0, 0), 'LEFT'),
            ('ALIGN', (1, 0), (1, 0), 'CENTER'),
            ('ALIGN', (2, 0), (2, 0), 'RIGHT'),
        ]))
        elements.append(header_table)
    else:
        elements.append(Paragraph('Reporte de Inventario Diario', title_style))

    elements.append(Paragraph('SenaFOOD - Sistema de Gestión', subtitle_style))
    elements.append(Paragraph(
        f'Fecha: {inv_diario.fecha}  |  '
        f'Responsable: {inv_diario.nombre_responsable}  |  '
        f'Generado el {datetime.now().strftime("%d/%m/%Y a las %H:%M")} '
        f'por {request.session.get("usuario_nombre", "Administrador")}',
        subtitle_style
    ))
    elements.append(Spacer(1, 6))

    # ── Línea verde ──
    elements.append(HRFlowable(width="100%", thickness=2, color=colors.HexColor('#28A745')))
    elements.append(Spacer(1, 8))

    # ── Cards estadísticas ──
    total        = inv_diario.total_productos
    diferencias  = inv_diario.total_diferencias
    sin_diff     = total - diferencias
    aumentos     = detalles.filter(diferencia__gt=0).count()
    disminucion  = detalles.filter(diferencia__lt=0).count()

    stats_data = [[
        Paragraph(f'<b>Total productos:</b> {total}',       stat_style),
        Paragraph(f'<b>Sin diferencias:</b> {sin_diff}',    stat_style),
        Paragraph(f'<b>Con aumento:</b> {aumentos}',        stat_style),
        Paragraph(f'<b>Con disminución:</b> {disminucion}', stat_style),
    ]]
    stats_table = Table(stats_data, colWidths=[185, 185, 185, 185])
    stats_table.setStyle(TableStyle([
        ('BACKGROUND',    (0,0), (-1,-1), colors.HexColor('#f0fff4')),
        ('GRID',          (0,0), (-1,-1), 0.5, colors.HexColor('#a8d5b5')),
        ('TOPPADDING',    (0,0), (-1,-1), 8),
        ('BOTTOMPADDING', (0,0), (-1,-1), 8),
        ('ALIGN',         (0,0), (-1,-1), 'CENTER'),
    ]))
    elements.append(stats_table)
    elements.append(Spacer(1, 12))

    # ── Observaciones ──
    if inv_diario.observaciones:
        obs_style = ParagraphStyle(
            'Obs', parent=styles['Normal'],
            fontSize=8, textColor=colors.HexColor('#155724'),
            backColor=colors.HexColor('#d4edda'),
            borderPadding=6, spaceAfter=10,
        )
        elements.append(Paragraph(f'Observaciones: {inv_diario.observaciones}', obs_style))
        elements.append(Spacer(1, 6))

    # ── Tabla detalle ──
    encabezados = [
        'N°', 'Producto', 'Categoría', 'Stock anterior',
        'Stock físico', 'Diferencia', 'Alerta mín.',
        'Estado anterior', 'Estado nuevo'
    ]
    data = [encabezados]

    for i, d in enumerate(detalles, 1):
        if d.diferencia > 0:
            dif_str = f'+{d.diferencia}'
        elif d.diferencia < 0:
            dif_str = str(d.diferencia)
        else:
            dif_str = '0'

        data.append([
            str(i),
            d.nombre_producto,
            d.categoria or '—',
            str(d.stock_anterior),
            str(d.stock_fisico),
            dif_str,
            str(d.alerta_minimos),
            d.estado_anterior or '—',
            d.estado_nuevo or '—',
        ])

    tabla = Table(data, repeatRows=1, colWidths=[25, 110, 80, 70, 70, 60, 60, 80, 80])
    tabla.setStyle(TableStyle([
        ('BACKGROUND',    (0,0), (-1,0), colors.HexColor('#28A745')),
        ('TEXTCOLOR',     (0,0), (-1,0), colors.white),
        ('FONTNAME',      (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE',      (0,0), (-1,0), 9),
        ('FONTSIZE',      (0,1), (-1,-1), 8),
        ('ALIGN',         (0,0), (-1,-1), 'CENTER'),
        ('VALIGN',        (0,0), (-1,-1), 'MIDDLE'),
        ('ROWBACKGROUNDS',(0,1), (-1,-1), [colors.white, colors.HexColor('#f0fff4')]),
        ('GRID',          (0,0), (-1,-1), 0.5, colors.HexColor('#a8d5b5')),
        ('TOPPADDING',    (0,0), (-1,-1), 6),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('LEFTPADDING',   (0,0), (-1,-1), 6),
        ('RIGHTPADDING',  (0,0), (-1,-1), 6),
    ]))
    elements.append(tabla)

    # ── Footer ──
    def add_footer(canvas, doc):
        canvas.saveState()
        canvas.setFont('Helvetica', 8)
        canvas.setFillColor(colors.HexColor('#888888'))
        canvas.drawString(30, 15, f'SenaFOOD - Inventario del {inv_diario.fecha}')
        canvas.drawRightString(landscape(A4)[0] - 30, 15, f'Página {doc.page}')
        canvas.setStrokeColor(colors.HexColor('#28A745'))
        canvas.setLineWidth(1)
        canvas.line(30, 25, landscape(A4)[0] - 30, 25)
        canvas.restoreState()

    doc.build(elements, onFirstPage=add_footer, onLaterPages=add_footer)
    return response


# ─────────────────────────────────────────────────────
# 6. EXPORTAR EXCEL — DETALLE DE UN INVENTARIO
# ─────────────────────────────────────────────────────

@sesion_requerida
def exportar_excel(request, pk):
    inv_diario = get_object_or_404(InventarioDiario, pk=pk)
    detalles   = inv_diario.detalles.all().order_by('nombre_producto')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'Inventario {inv_diario.fecha}'

    verde_fill  = PatternFill('solid', fgColor='28a745')
    blanco_font = Font(color='FFFFFF', bold=True)
    centro      = Alignment(horizontal='center', vertical='center')

    ws['A1'] = f'Inventario Diario — {inv_diario.fecha}'
    ws['A1'].font = Font(bold=True, size=13, color='28a745')
    ws['A2'] = f'Responsable: {inv_diario.nombre_responsable}'
    ws['A3'] = f'Observaciones: {inv_diario.observaciones or "—"}'
    ws.append([])

    encabezados = ['#', 'Producto', 'Categoría', 'Stock anterior',
                'Stock físico', 'Diferencia', 'Alerta mín.',
                'Costo unitario', 'Estado anterior', 'Estado nuevo']
    ws.append(encabezados)

    for cell in ws[5]:
        cell.fill      = verde_fill
        cell.font      = blanco_font
        cell.alignment = centro

    for i, d in enumerate(detalles, 1):
        ws.append([
            i,
            d.nombre_producto,
            d.categoria or '',
            d.stock_anterior,
            d.stock_fisico,
            d.diferencia,
            d.alerta_minimos,
            float(d.costo_unitario or 0),
            d.estado_anterior or '',
            d.estado_nuevo or '',
        ])

    anchos = [4, 28, 16, 14, 12, 12, 12, 15, 16, 14]
    for col, ancho in enumerate(anchos, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = ancho

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="inventario_{inv_diario.fecha}.xlsx"'}
    )

@sesion_requerida
def actualizar_alerta(request, id_producto):
    inv = get_object_or_404(Inventario, idproducto=id_producto)
    if request.method == 'POST':
        nuevo_valor = request.POST.get('alerta_minimos')
        try:
            inv.alerta_minimos = int(nuevo_valor)
            inv.update_at      = timezone.now()
            inv.save()
            messages.success(request, f'Alerta mínima actualizada a {nuevo_valor}.')
        except Exception as e:
            messages.error(request, f'Error: {str(e)}')
    return redirect('lista_inventario')